import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
root_str = str(ROOT)
if sys.path[:1] != [root_str]:
    if root_str in sys.path:
        sys.path.remove(root_str)
    sys.path.insert(0, root_str)

import streamlit as st
import importlib
import importlib.util

def _load_local_app():
    app_path = (ROOT / 'app.py').resolve()
    spec = importlib.util.spec_from_file_location('app', app_path)
    if spec is None or spec.loader is None:
        raise ImportError(f'Cannot load app.py from {app_path}')
    module = importlib.util.module_from_spec(spec)
    sys.modules['app'] = module
    spec.loader.exec_module(module)
    return module

try:
    import app as _app
    _app_file = getattr(_app, '__file__', '')
    if not _app_file or Path(_app_file).resolve() != (ROOT / 'app.py').resolve():
        _app = _load_local_app()
except Exception:
    _app = _load_local_app()

APP_TITLE = getattr(_app, 'APP_TITLE', 'PDF & Image Tools')
from app import *

st.set_page_config(page_title=f"{APP_TITLE} - PDF to EXCEL", layout="wide")
if st.button("← Home"):
    st.switch_page("app.py")

st.title("PDF to EXCEL")

deps = get_dependency_status()

st.info("Extracts tables is non-trivial; this tool exports page text into a spreadsheet (one sheet per page).")
f = st.file_uploader("PDF", type=["pdf"], key="pdf2xls")
if f and st.button("Convert", key="pdf2xls_btn"):
    import openpyxl

    reader = read_uploaded_pdf(f)
    wb = openpyxl.Workbook()
    for i, page in enumerate(reader.pages, start=1):
        if i == 1 and wb.active is not None:
            ws = wb.active
            ws.title = "Page 1"
        else:
            ws = wb.create_sheet(title=f"Page {i}")
        text = (page.extract_text() or "").splitlines()
        for r, line in enumerate(text, start=1):
            ws.cell(row=r, column=1, value=line)
    out = io.BytesIO()
    wb.save(out)
    st.download_button(
        "Download XLSX",
        out.getvalue(),
        file_name="document.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

